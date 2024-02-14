import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import csv
import openpyxl
from PyPDF2 import PdfFileReader

# Inicializar la lista de datos combinados
datos_combinados = []

# Función para cargar datos desde un archivo CSV o Excel
def cargar_datos(desde_excel=False):
    try:
        if desde_excel:
            datos = []
            libro_excel = openpyxl.load_workbook('datos.xlsx')
            hoja_excel = libro_excel.active
            for fila in hoja_excel.iter_rows(values_only=True):
                datos.append(tuple(fila))
        else:
            with open('datos.csv', newline='', encoding='utf-8') as archivo:
                lector_csv = csv.reader(archivo)
                datos = list(lector_csv)
        return datos
    except FileNotFoundError:
        return []

# Función para guardar datos en un archivo CSV o Excel
def guardar_datos(datos, a_excel=False):
    try:
        if a_excel:
            libro_excel = openpyxl.Workbook()
            hoja_excel = libro_excel.active
            for dato in datos:
                hoja_excel.append(dato)
            libro_excel.save('datos.xlsx')
        else:
            with open('datos.csv', 'w', newline='', encoding='utf-8') as archivo:
                escritor_csv = csv.writer(archivo)
                escritor_csv.writerows(datos)
    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar datos: {e}")

# Función para mostrar mensajes de confirmación
def mostrar_mensaje(titulo, mensaje):
    messagebox.showinfo(titulo, mensaje)

# Función para eliminar elemento seleccionado
def eliminar_elemento():
    seleccion = resultado.curselection()
    if seleccion:
        indice = seleccion[0]
        elemento_eliminar = datos_combinados[indice]
        confirmacion = messagebox.askokcancel("Eliminar", f"¿Estás seguro de eliminar: {elemento_eliminar[0]} - {elemento_eliminar[1]}?")
        if confirmacion:
            datos_combinados.pop(indice)
            guardar_datos(datos_combinados)
            mostrar_resultado(datos_combinados)
            mostrar_mensaje("Éxito", "Se eliminó el elemento.")
    else:
        mostrar_mensaje("Error", "Por favor, selecciona un elemento para eliminar.")

# Función para editar elemento seleccionado
def editar_elemento():
    seleccion = resultado.curselection()
    if seleccion:
        indice = seleccion[0]
        elemento_editar = datos_combinados[indice]

        # Crear ventana de edición
        ventana_edicion = tk.Toplevel(ventana)
        ventana_edicion.title("Editar Elemento")

        # Etiquetas y cuadros de entrada para edición
        nombre_label = ttk.Label(ventana_edicion, text="Nuevo Nombre:")
        correo_label = ttk.Label(ventana_edicion, text="Nueva Dirección de Correo:")
        nuevo_nombre_entry = ttk.Entry(ventana_edicion, width=30)
        nueva_correo_entry = ttk.Entry(ventana_edicion, width=30)

        # Configurar valores iniciales en los cuadros de entrada
        nuevo_nombre_entry.insert(0, elemento_editar[0])
        nueva_correo_entry.insert(0, elemento_editar[1])

        def aplicar_edicion():
            nuevo_nombre = nuevo_nombre_entry.get()
            nueva_direccion = nueva_correo_entry.get()

            if nuevo_nombre and nueva_direccion:
                datos_combinados[indice] = (nuevo_nombre, nueva_direccion)
                guardar_datos(datos_combinados)
                mostrar_resultado(datos_combinados)
                mostrar_mensaje("Éxito", "Se aplicaron los cambios.")
                ventana_edicion.destroy()
            else:
                mostrar_mensaje("Error", "Por favor, ingrese nuevo nombre y dirección de correo electrónico.")

        # Botón para aplicar la edición
        boton_aplicar_edicion = ttk.Button(ventana_edicion, text="Aplicar Edición", command=aplicar_edicion)

        # Colocar widgets en la ventana de edición
        nombre_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        correo_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        nuevo_nombre_entry.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        nueva_correo_entry.grid(row=1, column=1, padx=10, pady=5, sticky="w")
        boton_aplicar_edicion.grid(row=2, column=0, columnspan=2, pady=10)
    else:
        mostrar_mensaje("Error", "Por favor, selecciona un elemento para editar.")

# Función para cargar datos desde un archivo Excel
def cargar_datos_desde_excel():
    global datos_combinados
    datos_combinados = cargar_datos(desde_excel=True)
    mostrar_resultado(datos_combinados)
    mostrar_mensaje("Éxito", "Datos importados desde Excel.")

# Función para mostrar resultados en el cuadro de texto
def mostrar_resultado(lista):
    resultado.config(state=tk.NORMAL)
    resultado.delete(1.0, tk.END)
    for nombre, direccion in lista:
        resultado.insert(tk.END, f"{nombre}: {direccion}\n")
    resultado.config(state=tk.DISABLED)

# Función para mostrar todos los datos iniciales
def mostrar_resultado_inicial():
    mostrar_resultado(datos_combinados)

# Función para ordenar los datos alfabéticamente por nombre
def ordenar_por_nombre():
    datos_combinados_ordenados = sorted(datos_combinados, key=lambda x: x[0])
    mostrar_resultado(datos_combinados_ordenados)

# Función para ordenar los datos alfabéticamente por correo electrónico
def ordenar_por_correo():
    datos_combinados_ordenados = sorted(datos_combinados, key=lambda x: x[1])
    mostrar_resultado(datos_combinados_ordenados)

# Función para deshacer el orden y mostrar los datos originales
def deshacer_orden():
    mostrar_resultado(datos_combinados)

# Función para agregar un nuevo dato
def agregar_dato():
    nuevo_nombre = nombre_entry.get()
    nueva_direccion = correo_entry.get()

    if nuevo_nombre and nueva_direccion:
        datos_combinados.append((nuevo_nombre, nueva_direccion))
        guardar_datos(datos_combinados)
        mostrar_resultado(datos_combinados)
        mostrar_mensaje("Éxito", "Se agregó el nuevo dato.")
    else:
        mostrar_mensaje("Error", "Por favor, ingrese nuevo nombre y dirección de correo electrónico.")

# Función para abrir archivos PDF y Excel
def abrir_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf"), ("Excel files", "*.xlsx;*.xls")])
    if archivo:
        if archivo.endswith('.pdf'):
            try:
                with open(archivo, 'rb') as pdf_file:
                    pdf_reader = PdfFileReader(pdf_file)
                    num_paginas = pdf_reader.numPages
                mostrar_mensaje("Información", f"Se abrió el archivo PDF: {archivo}\nNúmero de páginas: {num_paginas}")
            except Exception as e:
                mostrar_mensaje("Error", f"No se pudo abrir el archivo PDF: {e}")
        elif archivo.endswith(('.xlsx', '.xls')):
            try:
                libro_excel = openpyxl.load_workbook(archivo)
                hojas_excel = libro_excel.sheetnames
                mostrar_mensaje("Información", f"Se abrió el archivo Excel: {archivo}\nHojas disponibles: {hojas_excel}")
            except Exception as e:
                mostrar_mensaje("Error", f"No se pudo abrir el archivo Excel: {e}")

# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Gestión de Nombres y Correos Electrónicos")

# Estilo ttk
estilo = ttk.Style()
estilo.configure("TButton", padding=5, font=('Helvetica', 10))

# Estilo ttk para el cuadro de texto
estilo_texto = ttk.Style()
estilo_texto.configure("TText", wrap="word")

# Botones
boton_mostrar_todo = ttk.Button(ventana, text="Mostrar Todo", command=mostrar_resultado_inicial)
boton_ordenar_correo = ttk.Button(ventana, text="Ordenar por Correo", command=ordenar_por_correo)
boton_ordenar_nombre = ttk.Button(ventana, text="Ordenar por Nombre", command=ordenar_por_nombre)
boton_deshacer = ttk.Button(ventana, text="Deshacer Orden", command=deshacer_orden)
boton_abrir_archivo = ttk.Button(ventana, text="Abrir Archivo", command=abrir_archivo)
boton_eliminar = ttk.Button(ventana, text="Eliminar Seleccionado", command=eliminar_elemento)
boton_editar = ttk.Button(ventana, text="Editar Seleccionado", command=editar_elemento)

# Etiquetas y cuadros de entrada
nombre_label = ttk.Label(ventana, text="Nombre:")
correo_label = ttk.Label(ventana, text="Correo Electrónico:")
nombre_entry = ttk.Entry(ventana)
correo_entry = ttk.Entry(ventana)

# Botón agregar
boton_agregar = ttk.Button(ventana, text="Agregar Dato", command=agregar_dato)

# Cuadro de texto para mostrar resultados
resultado = tk.Text(ventana, height=10, width=40, state=tk.DISABLED, font=('Helvetica', 10))
resultado.grid(row=1, column=0, columnspan=5, padx=10, pady=10, sticky="w")

# Scrollbar para el cuadro de texto
scrollbar = tk.Scrollbar(ventana, orient="vertical", command=resultado.yview)
resultado.config(yscrollcommand=scrollbar.set)
scrollbar.grid(row=1, column=5, pady=10, sticky="ns")

# Colocar widgets en la ventana
boton_mostrar_todo.grid(row=0, column=0, padx=10, pady=5, sticky="w")
boton_ordenar_correo.grid(row=0, column=1, padx=10, pady=5)
boton_ordenar_nombre.grid(row=0, column=2, padx=10, pady=5)
boton_deshacer.grid(row=0, column=3, padx=10, pady=5)
boton_abrir_archivo.grid(row=0, column=4, padx=10, pady=5)
boton_eliminar.grid(row=5, column=0, padx=10, pady=5, sticky="w")
boton_editar.grid(row=5, column=1, padx=10, pady=5)
nombre_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
correo_label.grid(row=3, column=0, padx=10, pady=5, sticky="w")
nombre_entry.grid(row=2, column=1, padx=10, pady=5, sticky="w")
correo_entry.grid(row=3, column=1, padx=10, pady=5, sticky="w")
boton_agregar.grid(row=4, column=0, columnspan=2, pady=5)

# Botón para exportar datos a un archivo Excel
boton_exportar_excel = ttk.Button(ventana, text="Exportar a Excel", command=lambda: guardar_datos(datos_combinados, a_excel=True))
boton_exportar_excel.grid(row=6, column=0, padx=10, pady=5, sticky="w")

# Botón para importar datos desde un archivo Excel
boton_importar_excel = ttk.Button(ventana, text="Importar desde Excel", command=lambda: cargar_datos_desde_excel())
boton_importar_excel.grid(row=6, column=1, padx=10, pady=5)

# Mostrar la ventana
ventana.mainloop()
